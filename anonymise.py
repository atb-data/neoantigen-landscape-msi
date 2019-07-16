from openpyxl import load_workbook

def parse_replacements(replacements):
  result = {}
  with open(replacements) as ref:
    for idx, line in enumerate(ref, 1):
      result[line.strip()] = f"HDCS{idx:04}"
  return result

def parse_all_replacements(paths):
  tumor_ids = []
  run_ids = []
  for path in paths:
    wb = load_workbook(path)
    for sheet in wb.worksheets:
      for row in sheet.iter_rows(
        sheet.min_row + 1, sheet.max_row,
        1, 1
      ):
        for cell in row:
          tumor_ids.append(str(cell.value))
      for row in sheet.iter_rows(
        sheet.min_row + 1, sheet.max_row,
        3, 3
      ):
        for cell in row:
          run_ids.append(str(cell.value))

  tumor_ids = sorted(list(set([id for id in tumor_ids if id is not None])))
  run_ids = sorted(list(set([id for id in run_ids if id is not None])))
  
  result = {}
  run_result = {}
  cc = 1
  ec = 1
  for tuid in tumor_ids:
    if tuid.startswith("MECA"):
      result[tuid] = f"HDES{ec:04}"
      ec += 1
    else:
      result[tuid] = f"HDCS{cc:04}"
      cc += 1

  for idx, runid in enumerate(run_ids, 1):
    run_result[runid] = str(idx)

  return result, run_result

def anonymise(infile, outfile, full_replacements, do_runs=True):
  replacements, run_replacements = full_replacements
  wb = load_workbook(infile)
  for sheet in wb.worksheets:
    tumor_ids = []
    replacement_ids = []
    for row in sheet.iter_rows(
      sheet.min_row + 1, sheet.max_row,
      1, 1
    ):
      for cell in row:
        tumor_ids.append(cell.value)
        if cell.value is not None:
          replacement_ids.append(replacements[cell.value])
          cell.value = replacements[cell.value]
    if do_runs:
      for idx, row in enumerate(sheet.iter_rows(
        sheet.min_row + 1, sheet.max_row,
        3, 3
      )):
        current_id = tumor_ids[idx]
        for cell in row:
          val = cell.value
          if val is not None:
            val = run_replacements[val]
            cell.value = val
  wb.save(outfile)