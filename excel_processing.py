from openpyxl import load_workbook
import pandas as pd

import random

def anonymized(tumor_list, prefix="CS"):
  """Anonymizes a list of tumor sample names."""
  result = {}
  alphabetical = sorted(tumor_list)
  for idx, elem in enumerate(alphabetical):
    result[elem] = elem
  return result

def tumor_characteristics(file, anon_output="outfiles/full_anonymization.csv"):
  """Extracts the tumor sample characteristics by sample name."""
  result = {}
  data = pd.read_csv(file, sep=";")
  names = list(data["TU_ID"])
  hereditary_status = list(data["MSI.type"])
  tumor_type = list(data["Tumor.type"])
  b2m_type = list(data["B2M.Seq"])
  anonymization = anonymized(names)
  anonymized_names = list(map(lambda x: anonymization[x], anonymization.keys()))
  with open(anon_output, "w") as csv_file:
    for idx in range(len(names)):
      csv_file.write(f"{names[idx]}{anonymized_names[idx]}\n")

      valid = True
      valid = valid and names[idx] != "#N/A"
      valid = valid and hereditary_status[idx] != "#N/A"
      valid = valid and tumor_type[idx] != "#N/A"
      valid = valid and b2m_type[idx] != "#N/A"

      item = {
        "name": names[idx],
        "anon": anonymized_names[idx],
        "hereditary": True if hereditary_status[idx] == "hereditary" else False,
        "type": True if tumor_type[idx] == "CRC" else False,
        "b2m": True if b2m_type[idx] == "mutated" else False,
      }

      if valid:
        result[names[idx]] = item
  
  return result

def parse_endometrium_mutations(csv_path):
  """Parses mutation rates from the endometrium raw data."""
  data = {}
  candidates = []
  with csv_path.open() as csv_file:
    header = next(csv_file)
    for line in csv_file:
      fields = line.split(";")
      tumor_id = fields[0]
      candidate = fields[1]
      print(tumor_id, fields[32], fields[33], fields[34])
      m2 = float(fields[32])
      m1 = float(fields[33])
      wt = float(fields[34])
      if not (tumor_id in data):
        data[tumor_id] = {}
      data[tumor_id][candidate] = {
        "wt": wt,
        "m1": m1,
        "m2": m2
      }
      if not (candidate in candidates):
        candidates.append(candidate)
  m1 = {}
  m2 = {}
  for key in data.keys():
    for candidate in candidates:
      if not (candidate + "_m1" in m1):
        m1[candidate + "_m1"] = 0.0
        m2[candidate + "_m2"] = 0.0
      if candidate in data[key]:
        m1[candidate + "_m1"] += data[key][candidate]["m1"]
        m2[candidate + "_m2"] += data[key][candidate]["m2"]
  for candidate in candidates:
    m1[candidate + "_m1"] /= len(data.keys())
    m2[candidate + "_m2"] /= len(data.keys())

  return {**m1, **m2}

def read_mutations(file, filter=lambda x: True):
  """Percentage of analyzed samples containing m1 frame and m2 frame peptides."""
  m1 = {}
  m2 = {}
  wb = load_workbook(file)
  for name in wb.sheetnames:
    sheet = wb[name]
    m1sum = 0.0
    count1 = 0
    for idx in range(1, len(sheet["AH"])):
      tumor_sample = sheet["A"][idx].value
      if filter(tumor_sample):
        val = sheet["AH"][idx].value
        if val == None:
          break
        count1 += 1
        m1sum += val
    m2sum = 0.0
    count2 = 0
    for idy in range(1, len(sheet["AG"])):
      tumor_sample = sheet["A"][idy].value
      if filter(tumor_sample):
        val = sheet["AG"][idy].value
        if val == None:
          break
        count2 += 1
        m2sum += val
    if count1 == 0:
      m1sum = 0.0
    else:
      m1sum /= count1

    if count2 == 0:
      m2sum = 0.0
    else:
      m2sum /= count2
    m1[name + "_m1"] = m1sum
    m2[name + "_m2"] = m2sum
  return m1, m2

def read_mutated_mutations(file, filter=lambda x: True):
  """Percentage of analyzed samples containing m1 frame and m2 frame peptides."""
  m1 = {}
  m2 = {}
  wb = load_workbook(file)
  for name in wb.sheetnames:
    sheet = wb[name]
    m1sum = 0.0
    count1 = 0
    for idx in range(1, len(sheet["AH"])):
      tumor_sample = sheet["A"][idx].value
      if filter(tumor_sample):
        val = sheet["AH"][idx].value
        if val == None:
          break
        count1 += 1 if val > 0.0 else 0
        m1sum += val
    m2sum = 0.0
    count2 = 0
    for idy in range(1, len(sheet["AG"])):
      tumor_sample = sheet["A"][idy].value
      if filter(tumor_sample):
        val = sheet["AG"][idy].value
        if val == None:
          break
        count2 += 1 if val > 0.0 else 0
        m2sum += val
    if count1 == 0:
      m1sum = 0.0
    else:
      m1sum /= count1

    if count2 == 0:
      m2sum = 0.0
    else:
      m2sum /= count2
    m1[name + "_m1"] = m1sum
    m2[name + "_m2"] = m2sum
  return m1, m2

def read_tumor_mutations(file, filter=lambda x: True):
  m1 = {}
  m2 = {}
  wb = load_workbook(file)
  for name in wb.sheetnames:
    m1[name + "_m1"] = {}
    m2[name + "_m2"] = {}
    sheet = wb[name]
    count1 = 0
    for idx in range(1, len(sheet["AH"])):
      tumor_sample = sheet["A"][idx].value
      if filter(tumor_sample):
        val = sheet["AH"][idx].value
        if val == None:
          break
        count1 += 1
        m1[name + "_m1"][tumor_sample] = val
    count2 = 0
    for idy in range(1, len(sheet["AG"])):
      tumor_sample = sheet["A"][idy].value
      if filter(tumor_sample):
        val = sheet["AG"][idy].value
        if val == None:
          break
        count2 += 1
        m2[name + "_m2"][tumor_sample] = val
  return m1, m2

def read_mutation_files(*files, method=read_mutations, filter=lambda x: True):
  """Reads mutation rates from files for m1 and m2 mutations."""
  fs = [
    method(f, filter=filter)
    for f in files
  ]
  result = {}
  for f in fs:
    result.update({**f[0], **f[1]})
  return result
