import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

from matplotlib import pyplot as plt
import matplotlib as mpl
from sklearn.decomposition import PCA, FastICA, KernelPCA
from sklearn.cluster import KMeans, SpectralClustering, AgglomerativeClustering, DBSCAN
from sklearn.svm import SVR, NuSVR, SVC
from sklearn.model_selection import train_test_split

from scipy.cluster.hierarchy import dendrogram, linkage
from scipy.cluster import hierarchy
import pandas as pd

def read_raw_tumor_data(path):
  wb = load_workbook(path)
  by_gene = {}
  for sheet in wb.sheetnames:
    by_tumor = {}
    first = True
    for row in wb[sheet].rows:
      if first:
        first = False
        continue
      tumor_id = row[0].value
      if tumor_id is not None:
        subrow = row[3:12]
        subrow = map(lambda x: float(x.value), subrow)
        by_tumor[tumor_id] = np.array(list(subrow))
    by_gene[sheet] = by_tumor

  return by_gene

def read_mutation_data(path, genes, tumors):
  with open(path) as csv:
    header = next(csv).strip().split(";")
    by_tumor = {}
    for line in csv:
      line = line.strip().split(";")
      tumor = line[0]
      heredity_status = (2 * float(line[1] == "hereditary") - 1) if line[1] != "" else 0.0
      b2m_status = (2 * float(line[12] == "mutated") - 1) if line[12] != "" else 0.5
      tnm = list(map(lambda x: float(x) if x != "#N/A" else 0, line[2:5]))
      by_tumor[tumor] = {
        "hereditary": heredity_status,
        "b2m": b2m_status,
        "tnm": tnm
      }
  hereditary_data = np.zeros((1, len(genes), len(tumors)))
  b2m_data = np.zeros((1, len(genes), len(tumors)))
  tnm_data = np.zeros((3, len(genes), len(tumors)))

  for idy, tumor in enumerate(tumors):
    if tumor in by_tumor:
      for idx, gene in enumerate(genes):
        hereditary_data[:, idx, idy] = by_tumor[tumor]["hereditary"]
        b2m_data[:, idx, idy] = by_tumor[tumor]["b2m"]
        tnm_data[:, idx, idy] = np.array(by_tumor[tumor]["tnm"])

  return hereditary_data, b2m_data, tnm_data

def read_gels_data(path, genes, tumors):
  strengths = {
    "strong": 0,
    "weak": 1,
    "garbage": 2
  }
  by_gene = {
    gene : np.zeros(3)
    for gene in genes
  }

  with open(path) as csv:
    header = next(csv).strip().split(";")
    for line in csv:
      line = line.strip().split(";")
      cohort = line[0]
      frame = int(line[2])
      gene = line[3].split("_")[0]
      if cohort == "European Caucasian" and frame == 1 and gene in genes:
        strength = strengths[line[1]]
        gels_50 = float(line[30])
        by_gene[gene][strength] = gels_50
  gels = np.zeros((3, len(genes), len(tumors)))
  for idx, gene in enumerate(genes):
    for idy, tumor in enumerate(tumors):
      gels[:, idx, idy] = by_gene[gene]

  nonzero_mean = gels[:, gels[2] != 0.0].mean(axis=1)
  for idx in range(len(genes)):
    for idy in range(len(tumors)):
      if gels[2, idx, idy] == 0.0:
        gels[:, idx, idy] = nonzero_mean
  return gels

def read_all_colon_data(path):
  p1 = path + "/QMR2.xlsx"
  p2 = path + "/QMR1.xlsx"
  by_gene_1 = read_raw_tumor_data(p1)
  by_gene_2 = read_raw_tumor_data(p2)

  genes_1 = [key for key in by_gene_1]
  genes_2 = [key for key in by_gene_2]

  tumors_1 = [
    tumor
    for gene in genes_1
    for tumor in by_gene_1[gene]
  ]
  tumors_2 = [
    tumor
    for gene in genes_2
    for tumor in by_gene_2[gene]
  ]
  genes = genes_1 + genes_2
  tumors = list(set(tumors_1 + tumors_2))

  nfeatures = 9
  data = np.zeros((nfeatures, len(genes), len(tumors)))

  for idx, gene in enumerate(genes):
    for idy, tumor in enumerate(tumors):
      if gene in by_gene_1 and tumor in by_gene_1[gene]:
        data[:, idx, idy] = by_gene_1[gene][tumor]
      elif gene in by_gene_2 and tumor in by_gene_2[gene]:
        data[:, idx, idy] = by_gene_2[gene][tumor]
      else:
        data[4, idx, idy] = 0.0

  return genes, tumors, data

def read_all_endometrium_data(path):
  path = path + "/ECQMR.xlsx"
  by_gene = read_raw_tumor_data(path)

  genes = [key for key in by_gene]

  tumors = [
    tumor
    for gene in genes
    for tumor in by_gene[gene]
  ]
  tumors = list(set(tumors))

  nfeatures = 9
  data = np.zeros((nfeatures, len(genes), len(tumors)))

  for idx, gene in enumerate(genes):
    for idy, tumor in enumerate(tumors):
      if gene in by_gene and tumor in by_gene[gene]:
        data[:, idx, idy] = by_gene[gene][tumor]
      else:
        data[4, idx, idy] = 0.0

  return genes, tumors, data

def cluster_and_plot(data, label_data, X=3, n_cluster=3, cluster_reduce=0, reduction=PCA, clustering=KMeans):
  fig, ax = plt.subplots(X, X, figsize=(30, 30))
  data_reduced = reduction(n_components=X).fit_transform(data)
  data_fit = data
  if cluster_reduce != 0:
    data_fit = reduction(n_components=cluster_reduce).fit_transform(data_fit)
  clustering = clustering(n_cluster).fit(data_fit)
  for idx in range(X):
    for idy in range(X):
      ax[idx, idy].scatter(
        data_reduced[:, idx], data_reduced[:, idy],
        c=clustering.labels_, cmap="tab10"
      )
      for i, label in enumerate(label_data):
        ax[idx, idy].annotate(label, (data_reduced[i, idx], data_reduced[i, idy]), size=4)
  fig.tight_layout()
  plt.savefig(f"cluster-plot-reduced-using-{reduction.__name__}-{X}-clustered-using-{clustering.__name__}-{n_cluster}-{cluster_reduce}.svg")
  return clustering

def mean_correct(data, genes, tumors):
  result = data.copy()
  nonzero_mean = result[:, data[4] != 0.0].mean(axis=1)
  for idx in range(len(genes)):
    for idy in range(len(tumors)):
      if result[4, idx, idy] == 0.0:
        result[:, idx, idy] = nonzero_mean
  return result

def nan_correct(data, genes, tumors):
  result = data.copy()
  for idx in range(len(genes)):
    for idy in range(len(tumors)):
      if result[4, idx, idy] == 0.0:
        result[:, idx, idy] = float("nan")
  return result

def frame_features(data):
  wt = np.expand_dims(data[1] + data[4] + data[7], axis=0)
  m1 = np.expand_dims(data[0] + data[3] + data[6], axis=0)
  m2 = np.expand_dims(data[8] + data[2] + data[5], axis=0)
  return np.concatenate((wt, m1, m2), axis=0)

def extract_views(data, genes, tumors):
  feature_size = data.shape[0]
  gene_size = data.shape[1]
  tumor_size = data.shape[2]
  data_genes = data.transpose(1, 2, 0).reshape(
    gene_size, tumor_size * feature_size
  )
  data_tumors = data.reshape(
    gene_size * feature_size, tumor_size
  ).transpose(1, 0)
  return data_genes, data_tumors

def parse_renamings(renamings, delimiter=","):
  """Reads all given internal to HUGO renamings from a csv file."""
  data = pd.read_csv(renamings, delimiter=delimiter)
  result = {}
  for row in data.itertuples():
    result[row[1]] = row[2]
  return result

def rename_candidate_hugo(candidate, renamings):
  """Renames a candidate name according to a renaming map."""
  name_expr = candidate.split(".")
  base_name = name_expr[0]
  if base_name in renamings:
    base_name = renamings[base_name]
  name_expr[0] = base_name
  result = ".".join(name_expr)
  return result

def rename_candidate(candidate, renamings=None):
  """Renames a candidate name according to a set of renamings."""
  result = candidate
  if renamings is not None:
    result = rename_candidate_hugo(candidate, renamings)
  if result == "RNF43_C6":
    result = "RNF43.A2"
  elif result == "RNF43_G7":
    result = "RNF43.A3"
  return result

def prepare_data(path):
  renamings = parse_renamings(path + "/Renamings.csv")

  crc_genes, crc_tumors, crc_data = read_all_colon_data(path)
  ec_genes, ec_tumors, ec_data = read_all_endometrium_data(path)

  crc_genes = list(map(
    lambda x: rename_candidate(x, renamings=renamings),
    crc_genes
  ))
  ec_genes = list(map(
    lambda x: rename_candidate(x, renamings=renamings),
    ec_genes
  ))

  hereditary, b2m, _ = read_mutation_data(path + "CRCQMR.csv", crc_genes, crc_tumors)
  gels = read_gels_data("/".join(path.split("/")[:-2]) + "/table_dump.csv", crc_genes, crc_tumors)
  gels = gels

  crc_data_m = mean_correct(crc_data, crc_genes, crc_tumors)
  crc_data_f = frame_features(crc_data_m)
  crc_data_hmf = frame_features(nan_correct(crc_data, crc_genes, crc_tumors))
  crc_data_mhb = np.concatenate((crc_data_m, hereditary, b2m), axis=0)
  crc_data_fhb = np.concatenate((crc_data_f, hereditary, b2m), axis=0)

  crc_data_fhb_0 = crc_data_fhb[:, :, crc_data_fhb[4, 0, :] == -1]
  crc_data_fhb_1 = crc_data_fhb[:, :, crc_data_fhb[4, 0, :] == 1]

  crc_data_fhb0 = crc_data_fhb_0[:3]
  crc_data_fhb1 = crc_data_fhb_1[:3]

  crc_data_fg = np.concatenate((crc_data_f, gels), axis=0)
  ec_data_m = mean_correct(ec_data, ec_genes, ec_tumors)
  ec_data_f = frame_features(ec_data_m)

  # common data:
  common_genes = np.array(list(set(crc_genes).intersection(set(ec_genes))))
  common_indices_crc = np.array([crc_genes.index(common) for common in common_genes])
  common_indices_ec = np.array([ec_genes.index(common) for common in common_genes])
  common_genes_order = sorted(list(range(len(common_genes))), key=lambda x: common_genes[x])
  ec_data_common = ec_data_f[:, common_indices_ec[common_genes_order], :]
  crc_data_common = crc_data_f[:, common_indices_crc[common_genes_order], :]
  common_data = np.concatenate((crc_data_common, ec_data_common), axis=2)
  common_genes = common_genes[common_genes_order]
  common_tumors = crc_tumors + ec_tumors

  label_dict = {
    "common": {
      "tumor": common_tumors,
      "gene": common_genes
    },
    "crc": {
      "tumor": crc_tumors,
      "gene": crc_genes
    },
    "ec": {
      "tumor": ec_tumors,
      "gene": ec_genes
    }
  }

  data_dict = {
    key : {
      "tumor": {},
      "gene": {}
    }
    for key in ("common", "crc", "ec")
  }

  data_dict["crc"]["gene"]["gels"] = gels[1, :, 0]

  data_dict["common"]["gene"]["f"], data_dict["common"]["tumor"]["f"] = \
    extract_views(common_data, common_genes, common_tumors)

  data_dict["crc"]["gene"]["m"], data_dict["crc"]["tumor"]["m"] = \
    extract_views(crc_data_m, crc_genes, crc_tumors)
  data_dict["crc"]["gene"]["f"], data_dict["crc"]["tumor"]["f"] = \
    extract_views(crc_data_f, crc_genes, crc_tumors)
  data_dict["crc"]["gene"]["1-wt"], data_dict["crc"]["tumor"]["1-wt"] = \
    extract_views(1.0 - crc_data_f[0:1], crc_genes, crc_tumors)
  data_dict["crc"]["gene"]["hm-1-wt"], data_dict["crc"]["tumor"]["hm-1-wt"] = \
    extract_views(1.0 - crc_data_hmf[0:1], crc_genes, crc_tumors)
  data_dict["crc"]["gene"]["fhb0"], data_dict["crc"]["tumor"]["fhb0"] = \
    extract_views(crc_data_fhb0, crc_genes, crc_tumors)
  data_dict["crc"]["gene"]["fhb1"], data_dict["crc"]["tumor"]["fhb1"] = \
    extract_views(crc_data_fhb1, crc_genes, crc_tumors)
  data_dict["crc"]["gene"]["mhb"], data_dict["crc"]["tumor"]["mhb"] = \
    extract_views(crc_data_mhb, crc_genes, crc_tumors)
  data_dict["crc"]["gene"]["fg"], data_dict["crc"]["tumor"]["fg"] = \
    extract_views(crc_data_fg, crc_genes, crc_tumors)
  data_dict["ec"]["gene"]["m"], data_dict["ec"]["tumor"]["m"] = \
    extract_views(ec_data_m, ec_genes, ec_tumors)
  data_dict["ec"]["gene"]["f"], data_dict["ec"]["tumor"]["f"] = \
    extract_views(ec_data_f, ec_genes, ec_tumors)
  data_dict["ec"]["gene"]["1-wt"], data_dict["ec"]["tumor"]["1-wt"] = \
    extract_views(1.0 - ec_data_f[0:1], ec_genes, ec_tumors)

  return label_dict, data_dict

def analysis_name(analysis):
  kind, head, args, kwargs, name = analysis
  args_string = ",".join(map(lambda x: f"{x}", args))
  kwargs_string = ",".join(map(lambda x: f"{x}:{kwargs[x]}", kwargs))
  return f"{kind}-{head.__name__}-args-({args_string})-kwargs-({kwargs_string})"

def perform_analysis(data, analysis):
  kind, head, args, kwargs, name = analysis
  if kind == "pca":
    return head(*args, **kwargs).fit_transform(data)
  if kind in ("cluster", "bicluster"):
    return head(*args, **kwargs).fit(data)
  assert False

def analyse(data, analysis_list):
  result = {}
  for tumor_type in data:
    result[tumor_type] = {}
    for data_view in data[tumor_type]:
      result[tumor_type][data_view] = {}
      for feature_set in data[tumor_type][data_view]:
        if feature_set not in ["gels", "hm-1-wt"]:
          result[tumor_type][data_view][feature_set] = {}
          for index, analysis in enumerate(analysis_list):
            input_data = data[tumor_type][data_view][feature_set]
            result[tumor_type][data_view][feature_set][analysis[-1]] = \
              perform_analysis(input_data, analysis)
  return result

def plot_result(clustering, pca_result, data, labels, keys, gels=None, endometrium_colon_color=True):
  tumor_type, data_view, feature_set, analysis, pca = keys
  data_array = data[tumor_type][data_view][feature_set]
  data_reduced = pca_result
  label_data = labels[tumor_type][data_view]

  c = clustering
  cmap = "tab10"

  if endometrium_colon_color and tumor_type == "common":
    c = [0 for _ in labels["crc"]["tumor"]] + [1 for _ in labels["ec"]["tumor"]]

  X = data_reduced.shape[1]
  fig, ax = plt.subplots(X, X, figsize=(30, 30))
  for idx in range(X):
    for idy in range(X):
      ax[idx, idy].scatter(data_reduced[:, idx], data_reduced[:, idy], c=c, cmap=cmap)
      for i, label in enumerate(label_data):
        ax[idx, idy].annotate(label, (data_reduced[i, idx], data_reduced[i, idy]), size=4)
  fig.tight_layout()
  plt.savefig(f"cluster-plot-{tumor_type}-{data_view}-{feature_set}-{pca}-{analysis}.svg")
  plt.close("all")
  return clustering

def plot_dendrograms(data, labels):
  cmap = mpl.cm.get_cmap("tab10")
  hierarchy.set_link_color_palette([
    mpl.colors.to_hex(cmap(idx)) for idx in range(10)
  ])
  linkage_gene_total = linkage(
    data_dict["crc"]["gene"]["f"], 'ward',
    optimal_ordering=True
  )
  linkage_gene_0 = linkage(
    data_dict["crc"]["gene"]["fhb0"], 'ward',
    optimal_ordering=True
  )
  linkage_gene_1 = linkage(
    data_dict["crc"]["gene"]["fhb1"], 'ward',
    optimal_ordering=True
  )
  linkage_tumor = linkage(
    data_dict["common"]["tumor"]["f"], 'ward',
    optimal_ordering=True
  )

  fig, ax = plt.subplots(figsize=(5, 10))
  dendrogram(linkage_gene_total,
             orientation='left',
             above_threshold_color="grey",
             color_threshold=linkage_gene_total[-2, 2],
             labels=labels["crc"]["gene"],
             show_leaf_counts=True,
             ax=ax)
  ax.axvline(linkage_gene_total[-2, 2], color="grey", linestyle="--")

  plt.savefig("dendrogram-total.svg", dpi=600)

  fig, ax = plt.subplots(figsize=(5, 10))
  dendrogram(linkage_gene_0,
             orientation='left',
             above_threshold_color="grey",
             color_threshold=linkage_gene_0[-2, 2],
             labels=labels["crc"]["gene"],
             show_leaf_counts=True,
             ax=ax)
  ax.axvline(linkage_gene_0[-2, 2], color="grey", linestyle="--")
  ax.set_xlim(12, 0)

  plt.savefig("dendrogram-0.svg", dpi=600)

  fig, ax = plt.subplots(figsize=(5, 10))
  dendrogram(linkage_gene_1,
             orientation='left',
             above_threshold_color="grey",
             color_threshold=linkage_gene_0[-2, 2],
             labels=labels["crc"]["gene"],
             show_leaf_counts=True,
             ax=ax)
  ax.axvline(linkage_gene_0[-2, 2], color="grey", linestyle="--")
  ax.set_xlim(12, 0)

  plt.savefig("dendrogram-1.svg", dpi=600)

  fig, ax = plt.subplots(figsize=(5, 10))
  dendrogram(linkage_tumor,
             orientation='left',
             above_threshold_color="grey",
             color_threshold=linkage_tumor[-1, 1],
             labels=["CRC" for _ in labels["crc"]["tumor"]] + ["EC" for _ in labels["ec"]["tumor"]],
             show_leaf_counts=True,
             ax=ax)

  plt.savefig("dendrogram-t.svg", dpi=600)

def plot_clusters(data, labels, features="f"):
  gene_data = data["crc"]["gene"][features]
  gels_data = data["crc"]["gene"]["gels"]
  gene_labels = labels["crc"]["gene"]

  clustering = AgglomerativeClustering(n_clusters=3).fit(gene_data)
  reduced = KernelPCA(n_components=2, kernel="rbf", gamma=0.066).fit_transform(gene_data)

  cluster_indices = [clustering.labels_ == idx for idx in range(3)]
  names = [
    np.array(gene_labels)[index]
    for index in cluster_indices
  ]
  print("NAMES:", names)
  gels_mean = [
    gels_data[index].mean()
    for index in cluster_indices
  ]
  print("MEAN:", gels_mean)
  gels_std = [
    gels_data[index].std()
    for index in cluster_indices
  ]
  print("STD:", gels_std)

  fig, ax = plt.subplots(figsize=(10, 10))
  sc = ax.scatter(reduced[:, 0], reduced[:, 1], c=gels_data, cmap="Reds")
  for i, label in enumerate(gene_labels):
    ax.annotate(label, (reduced[i, 0] + 0.005, reduced[i, 1] + 0.005), size=8)

  plt.savefig(f"clustering-{features}.svg", dpi=600)

def plot_endometrium(data, labels):
  tumor_data = data["common"]["tumor"]["f"]
  tumor_labels = labels["common"]["tumor"]

  reduced = PCA(n_components=2).fit_transform(tumor_data)

  c = [0 for _ in labels["crc"]["tumor"]] + [1 for _ in labels["ec"]["tumor"]]

  fig, ax = plt.subplots(figsize=(10, 10))
  sc = ax.scatter(reduced[:, 0], reduced[:, 1], c=c, cmap="tab10")

  plt.savefig("crc-ec.svg", dpi=600)

def result_to_data(analysis_result):
  data = {}
  for tumor_type in analysis_result:
    data[tumor_type] = {}
    for data_view in analysis_result[tumor_type]:
      data[tumor_type][data_view] = {}
      for feature_set in analysis_result[tumor_type][data_view]:
        for key in analysis_result[tumor_type][data_view][feature_set]:
          data[tumor_type][data_view][feature_set] = \
            analysis_result[tumor_type][data_view][feature_set][key]
  return data

if __name__ == "__main__":
  import os

  path = os.path.dirname(os.path.realpath(__file__))
  path = path + "/testfiles/"
  label_dict, data_dict = prepare_data(path)

  pca_result = analyse(data_dict, [
    (
      "pca",
      KernelPCA, [],
      {
        "n_components": ncomp,
        "kernel": "rbf",
        "gamma": gamma
      },
      f"RBF-{gamma}-PCA-{ncomp}"
    )
    for ncomp in range(4, 5)
    for gamma in [0.066]
  ])
  cluster_result = analyse(data_dict, [
    (
      "cluster",
      AgglomerativeClustering, [],
      {"n_clusters": nclust},
      f"Agglomerative-{nclust}"
    )
    for nclust in range(3, 4)
  ])

  plot_endometrium(data_dict, label_dict)
  plot_clusters(data_dict, label_dict)
  plot_clusters(data_dict, label_dict, features="fhb0")
  plot_clusters(data_dict, label_dict, features="fhb1")
  plot_dendrograms(data_dict, label_dict)

  for tumor_type in cluster_result:
    for data_view in cluster_result[tumor_type]:
      for feature_set in cluster_result[tumor_type][data_view]:
        for analysis in cluster_result[tumor_type][data_view][feature_set]:
          clustering_result = cluster_result[tumor_type][data_view][feature_set][analysis]
          pcas = pca_result[tumor_type][data_view][feature_set]
          labels = clustering_result.labels_
          for pca in pcas:
            one_pca = pcas[pca]
            plot_result(
              labels, one_pca, data_dict, label_dict,
              (tumor_type, data_view, feature_set, analysis, pca),
              gels=data_dict["crc"]["gene"]["gels"]
            )
